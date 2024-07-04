import pdfplumber
import tabula
import fitz
import re
import time
import tabula
import pandas as pd
import pdfplumber
# from .get_similarity import compare_dictionaries
from .get_similarity import compare_dictionaries
import openpyxl


def get_standard_document_as_dict(wb, sheet_name):
    """
    获取吉森标准ce表的表格红色参数信息
    :param standard_excel: 吉森ce表
    :param sheet_name: excel表内的工作表
    :return: 一个字典
    """
    # Load the Excel file
    # wb = openpyxl.load_workbook(standard_excel)
    # Access the specified sheet
    sheet = wb[sheet_name]

    red_text_dict = {}  # Create a dictionary to store the results

    # Traverse each row
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        # Assume the table's first column is not empty and starts from the second column
        table_start_column = None
        for cell in row:
            if cell.value is not None and table_start_column is None:
                table_start_column = cell.column

        # If the start of the table is found
        if table_start_column is not None:
            # Get the value of the first column of the table
            first_cell_value = row[table_start_column - 1].value
            red_texts = []
            for cell in row[table_start_column:]:
                if cell.font and cell.font.color and cell.font.color.rgb == 'FFFF0000':  # Look for red font
                    # Add the red font value to the list
                    red_texts.append(cell.value)

            if red_texts:
                # If the key is already in the dictionary, append the new red text value
                if first_cell_value in red_text_dict:
                    red_text_dict[first_cell_value].extend(red_texts)
                else:
                    # Otherwise, create a new key in the dictionary
                    red_text_dict[first_cell_value] = red_texts

    # print(red_text_dict)
    # Remove None values from the dictionary
    for key in list(red_text_dict.keys()):
        red_text_dict[key] = [value for value in red_text_dict[key] if value is not None]
        # Remove keys with empty lists
        if not red_text_dict[key]:
            del red_text_dict[key]

    red_text_dict = update_key_standard_dict(red_text_dict, wb, sheet_name)
    return red_text_dict


def update_key_standard_dict(data_dict, wb, sheet_name):
    """
    对get_standard_document_as_dict函数获取到字典再处理下，
    将字典中值为xxxx-xx的键设为CE-sign。
    如果没有匹配的值，则从指定的Excel表中提取所有文本，找到匹配的内容来更新data_dict。

    :param data_dict: 一个字典
    :param wb: Excel 工作簿对象
    :param sheet_name: 工作表名称
    :return: 修改后的字典
    """
    # 定义匹配'xxxx-xx'格式的正则表达式
    ce_sign_pattern = re.compile(r'\b\d{4}-\d{2}\b')

    # 创建一个新字典用于存储更新后的结果
    updated_dict = {}
    ce_sign_found = False

    # 遍历原始字典中的项，检查是否存在符合模式的值
    for key, values in data_dict.items():
        values[0] = str(values[0])
        if values and ce_sign_pattern.match(values[0]):
            updated_dict['CE-sign'] = values
            ce_sign_found = True
        else:
            updated_dict[key] = values

    # 如果没有在原始字典中找到符合模式的值，则从Excel表中提取所有文本
    if not ce_sign_found:
        sheet = wb[sheet_name]
        matching_texts = [
            ce_sign_pattern.search(str(cell.value)).group() for row in sheet.iter_rows()
            for cell in row if cell.value and ce_sign_pattern.search(str(cell.value))
        ]
        if matching_texts:
            updated_dict['CE-sign'] = matching_texts

    return updated_dict


def extract_with_pdfplumber(pdf_path):
    """
    使用 pdfplumber 从 PDF 的第一页提取最大的表格。
    :param pdf_path: PDF 文件的路径
    :return: 包含表格数据的字典
    """
    with pdfplumber.open(pdf_path) as pdf:
        first_page = pdf.pages[0]  # 获取第一页
        tables = first_page.extract_tables()  # 提取所有表格

        if not tables:
            raise ValueError("No tables found with pdfplumber")

        # 获取最大的表格
        largest_table = max(tables, key=lambda table: len(table))
        num_rows = len(largest_table)
        num_columns = len(largest_table[0]) if num_rows > 0 else 0

        table_dict = {}
        for row in largest_table[1:]:  # 假设第一行是表头，从第二行开始处理数据
            if row[0] is not None:  # 确保键不为None
                key = row[0].strip()  # 删除可能的前后空白字符
                values = [item for item in row[1:] if item is not None]  # 过滤掉None值
                table_dict[key] = values

        return table_dict


def extract_with_tabula(pdf_path):
    """
    使用 tabula-py 从 PDF 的第一页提取最大的表格。
    :param pdf_path: PDF 文件的路径
    :return: 包含表格数据的字典
    """
    tables = tabula.read_pdf(pdf_path, pages=1, multiple_tables=True)

    if not tables or all(table.empty for table in tables):
        raise ValueError("No tables found with tabula")

    # 获取最大的表格
    largest_table = max(tables, key=lambda t: len(t))
    num_rows, num_columns = largest_table.shape

    table_dict = {}
    for index, row in largest_table.iterrows():  # 遍历表格的每一行
        if row[0]:  # 确保键不为空
            key = str(row[0]).strip()  # 删除可能的前后空白字符
            values = [str(item).strip() for item in row[1:] if
                      item is not None and str(item).strip().lower() != 'nan']  # 过滤掉 None 值和 'nan'
            if key.lower() != 'nan' and values:  # 确保键和值不为 'nan' 并且值列表不为空
                table_dict[key] = values

    return table_dict


def extract_table_from_pdf(pdf_path):
    """
    先尝试使用 pdfplumber 提取表格，如果失败或者提取的表格数量小于等于5，则使用 tabula 提取。
    :param pdf_path: PDF 文件的路径
    :return: 包含表格数据的字典
    """
    try:
        table_dict = extract_with_pdfplumber(pdf_path)
        if len(table_dict) <= 5:
            raise ValueError("用pdfplumber获取到的键值对小于6个")
        print("Extracted using pdfplumber:")
        return table_dict

    except Exception as e:
        print(f"Pdfplumber extraction failed or insufficient data: {e}")
        try:
            table_dict = extract_with_tabula(pdf_path)
            print("Extracted using tabula:")
            return table_dict
        except Exception as e:
            print(f"Tabula extraction failed: {e}")
            return {}


def remove_empty_lists(input_dict):
    """
    删除字典中值列表为空的键值对，并返回一个新的字典。

    :param input_dict: 要处理的字典
    :return: 更新后的字典，其中不包含空列表的键值对
    """
    # 使用字典推导式来创建一个新的字典，其中仅包含非空列表的键值对
    new_dict = {key: [item for item in value if item] for key, value in input_dict.items() if
                value and any(item for item in value if item)}

    return new_dict


def add_ce_signs_to_dict(doc, input_dict):
    """
    从PDF文件中读取文本，查找所有符合XXXX-XX格式的字符串（X是1到9的数字），
    并将它们作为列表添加到字典中，键为'CE-sign'。

    :param pdf_path: PDF文件的路径
    :param input_dict: 要更新的字典
    :return: None（原地修改字典）
    """

    # 用于存储所有找到的CE标记的列表
    ce_signs = []

    # 定义正则表达式
    pattern = re.compile(r'\b[0-9]{4}[-/][0-9]{2}\b')

    # 遍历PDF中的每一页
    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        text = page.get_text()

        # 在当前页的文本中查找所有匹配的字符串
        matches = pattern.findall(text)
        ce_signs.extend(matches)
    # 如果找到了匹配的字符串，则更新输入字典
    if ce_signs:
        input_dict['CE-sign'] = ce_signs
    return input_dict


def compare(wb, work_table, doc, PDF_PATH1):

    red_text_data = get_standard_document_as_dict(wb, work_table)
    print(f"excel表格:{red_text_data}")

    table_data = extract_table_from_pdf(PDF_PATH1)
    table_data = remove_empty_lists(table_data)
    table_data = add_ce_signs_to_dict(doc, table_data)
    print(f"pdf表格:{table_data}")
    message_dict = compare_dictionaries(red_text_data, table_data)
    return message_dict




