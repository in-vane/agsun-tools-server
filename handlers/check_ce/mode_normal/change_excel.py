import base64
from io import BytesIO
from PIL import Image
import fitz
import openpyxl
from jpype import JClass
from openpyxl.styles import Border, Side
import subprocess
import os
import shutil
import tempfile
import jpype

from main import MainHandler
import tornado
from tornado.concurrent import run_on_executor
from config import CONTENT_TYPE_PDF, LIBREOFFICE_PATH
from utils import convert_files_to_bytesio, ensure_directory_exists

import asposecells
jpype.startJVM()
from asposecells.api import Workbook, FileFormatType, PdfSaveOptions

from .get_table_message import compare
from save_filesys_db import save_ce


# LIBREOFFICE_PATH = "/usr/bin/soffice"
CODE_SUCCESS = 0
CODE_ERROR = 1
EXCEL_PATH = './assets/excel/temp.xlsx'
PDF_PATH = './assets/pdf/temp.pdf'
BASE64_PNG = 'data:image/png;base64,'


# 将错误的地方在excel文件框标出
def change_excel(wb, work_table, message_dict):
    """
    该函数主要是为了，呈现错误信息到吉盛标准ce表上
    :param excel: 吉森标准ce表
    :param work_table: ce表中的工作表
    :param message_dict: 错误信息
    :return:
    """
    # 加载工作簿并选择工作表
    # wb = load_workbook(excel)
    sheet = wb[work_table]

    # 用绿色定义边框样式
    # 用绿色定义边框样式
    green_border = Border(
        left=Side(style='thick', color='0000FF'),
        right=Side(style='thick', color='0000FF'),
        top=Side(style='thick', color='0000FF'),
        bottom=Side(style='thick', color='0000FF')
    )
# 检查 message_dict 中是否包含 'CE-sign' 键
    if 'CE-sign' in message_dict:
        ce_values = message_dict['CE-sign']
        print(ce_values)
        if ce_values:
            ce_value = ce_values[0]  # 获取'CE-sign'对应列表的第一个元素
            # 遍历工作表中的每一行和每一个单元格
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and str(cell.value) == str(ce_value):  # 如果单元格的值等于 ce_value
                        cell.border = green_border  # 更新边框为绿色

    # 遍历工作表中的每一行，更新其他指定的值的边框
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in message_dict and cell.value != 'CE-sign':  # 排除'CE-sign'键
                red_texts_positions = message_dict[cell.value]  # 获取红色文本位置的列表
                red_texts_count = 0  # 找到的红色文本单元格的计数器

                for row_cell in row:
                    # 检查单元格的字体是否为红色
                    if row_cell.font and row_cell.font.color and row_cell.font.color.type == 'rgb' and row_cell.font.color.value == 'FFFF0000':
                        red_texts_count += 1  # 递增计数器
                        # 如果当前计数在位置列表中，则更新边界
                        if red_texts_count in red_texts_positions:
                            row_cell.border = green_border

    # 保存文件
    wb.save(EXCEL_PATH)


def excel_to_iamge(excel_path, num, zoom=2.0):
    # fitz 的索引是从0开始，而用户是从1开始
    num = num - 1
    # 构建转换命令
    command = [
        LIBREOFFICE_PATH,
        "--headless",
        "--convert-to", "xlsx",
        'vvv',
        "--outdir", '.',
        excel_path
    ]
    # 执行转换命令
    try:
        subprocess.run(command, shell=False, check=True)
        print(f"The file has been converted ")
    except subprocess.CalledProcessError as e:
        print(f"An error occurred: {e}")
    file_name = os.path.splitext(
        os.path.basename(excel_path))[0]
    output_pdf_path = file_name + ".pdf"
    output_xlsx_path = file_name + ".xlsx"
    workbook = Workbook(output_xlsx_path)
    saveOptions = PdfSaveOptions()
    saveOptions.setOnePagePerSheet(True)

    workbook.save(output_pdf_path, saveOptions)
    # 使用fitz (PyMuPDF)将PDF文件的特定页面转换为Base64编码的图片字符串
    doc = fitz.open(output_pdf_path)
    if num < len(doc):
        page = doc.load_page(num)
        # 设置变换矩阵以提高分辨率
        matrix = fitz.Matrix(zoom, zoom)
        # 将页面转换为高分辨率图像
        pix = page.get_pixmap(matrix=matrix)
        img_bytes = pix.tobytes("png")
        img_base64 = base64.b64encode(img_bytes).decode("utf-8")
    else:
        img_base64 = None
        print("Specified page number exceeds the PDF's page count.")

    doc.close()
    # 删除生成的PDF文件
    os.remove(output_pdf_path)
    os.remove(output_xlsx_path)
    # jpype.shutdownJVM()
    return f"{BASE64_PNG}{img_base64}"

def pdf_to_image(doc, zoom=2.0):
    # 加载第一页
    page = doc.load_page(0)  # 页面索引从0开始，0表示第一页
    # 设置变换矩阵以提高分辨率
    matrix = fitz.Matrix(zoom, zoom)
    # 将页面转换为高分辨率图像
    pix = page.get_pixmap(matrix=matrix)
    # 将图像数据转换为PNG格式的二进制数据
    img_data = pix.tobytes("png")
    # 对图像数据进行Base64编码
    img_base64 = base64.b64encode(img_data).decode('utf-8')
    return f"{BASE64_PNG}{img_base64}"

# 将xls文件转化为xlsx文件
def convert_xls_bytes_to_xlsx(file_bytes):
    """
    使用LibreOffice将xls文件的字节流转换为xlsx格式的字节流。
    :param file_bytes: 要转换的xls文件的字节流。
    :return: 转换后的xlsx文件的字节流。
    """
    # 创建一个临时文件来保存字节流
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xls') as tmp_file:
        tmp_file_name = tmp_file.name
        tmp_file.write(file_bytes)
        tmp_file.flush()

    # 获取当前目录作为输出目录
    output_dir = tempfile.mkdtemp()

    try:
        # 构建转换命令
        cmd = [
            LIBREOFFICE_PATH,
            "--headless",
            "--convert-to",
            "xlsx:Calc MS Excel 2007 XML",
            tmp_file_name,
            "--outdir",
            output_dir
        ]
        # 调用命令行执行转换
        subprocess.run(cmd, check=True)
        # 生成的.xlsx文件路径
        generated_file_path = os.path.join(output_dir, os.path.splitext(
            os.path.basename(tmp_file_name))[0] + ".xlsx")
        # 读取生成的.xlsx文件的字节流
        with open(generated_file_path, 'rb') as f:
            xlsx_bytes = f.read()
        return xlsx_bytes

    except subprocess.CalledProcessError as e:
        print(f"转换失败: {e}")
        return None
    except Exception as e:
        print(f"发生错误: {e}")
        return None
    finally:
        # 清理临时文件和目录
        os.remove(tmp_file_name)
        shutil.rmtree(output_dir)


# 判断文件是那种excel文件
def determine_file_type(excel_bytes):
    # 将前几个字节转换为十六进制表示
    file_header = excel_bytes[:8].hex()
    # 判断文件类型
    if file_header.startswith('d0cf11e0a1b11ae1'):
        return 'xls'
    elif file_header.startswith('504b0304'):
        return 'xlsx'
    else:
        return 'Unknown'


def checkTags(username, excel_file, pdf_file, num, file_extension, excel):
    ensure_directory_exists(EXCEL_PATH)
    ensure_directory_exists(PDF_PATH)
    if file_extension == 'xls':
        print("这是xls文件")
        excel_file = convert_xls_bytes_to_xlsx(excel_file)
        print("转换为xlsx文件成功")
    doc = fitz.open(pdf_file)
    wb = openpyxl.load_workbook(filename=BytesIO(excel_file))
    sheet_names = wb.sheetnames
    print(f"一共工作表:{sheet_names}")
    if sheet_names[0] != 'label':
        work_table = sheet_names[num - 1]
    else:
        work_table = sheet_names[num]
    print(f"工作表为: {work_table}")
    message_dict = compare(wb, work_table, doc, pdf_file)
    change_excel(wb, work_table, message_dict)
    excel_image_base64 = excel_to_iamge(EXCEL_PATH, num)
    os.remove(EXCEL_PATH)
    pdf_image_base64 = pdf_to_image(doc)
    doc.close()
    wb.close()
    save_ce(username['username'], CODE_SUCCESS, pdf_file, excel, excel_image_base64, pdf_image_base64)
    return CODE_SUCCESS, excel_image_base64, pdf_image_base64, None


# 测试
# def pdf_to_bytes(file_path):
#  with open(file_path, 'rb') as file:
#     bytes_content = file.read()
# return bytes_content
class CEHandler(MainHandler):
    @run_on_executor
    def process_async(self, username, excel_file, pdf_file, num, file_extension, excel):
        return checkTags(username, excel_file, pdf_file, num, file_extension, excel)

    async def post(self):
        username = self.current_user
        params = tornado.escape.json_decode(self.request.body)
        mode = params['mode']
        num = int(params['sheet'])
        file_paths = params["filePath"]
        for file_path in file_paths:
            ext = os.path.splitext(file_path)[1].lower()
            if ext in ['.xlsx', '.xls']:
                excel = file_path
            elif ext == '.pdf':
                file_pdf = file_path
        # 获取文件后缀名
        file_extension = os.path.splitext(excel)[1].lstrip('.')
        file_excel = convert_files_to_bytesio(excel)
        if mode == 0:
            code, excel_image_base64, pdf_image_base64, msg = await self.process_async(username,
                                                                    file_excel, file_pdf, num, file_extension, excel)
        custom_data = {
            "code": code,
            "data": {
                "excel_image_base64": excel_image_base64,
                "pdf_image_base64": pdf_image_base64
            },
            "msg": msg
        }
        self.write(custom_data)