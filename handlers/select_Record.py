from main import MainHandler
import tornado
from tornado.concurrent import run_on_executor
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

sys.path.append("..")

CODE_SUCCESS = 0
CODE_ERROR = 1
from utils import add_url

from config import RECORD_FILE_ROOT, IMAGE_ROOT, FLIES_ROOT, FRONT
from model import record_count
import os
from datetime import datetime

def generate_excel_report(data, filename=None):
    """
    生成Excel报表并返回保存地址，同时显示运行结束时间和耗时。

    参数:
    data (dict): 用户数据字典。
    output_file (str): 保存的Excel文件名，默认为'报表.xlsx'。

    返回:
    str: Excel文件的保存路径。
    """
    # 确保 RECORD_FILE_ROOT 目录存在，如果不存在则创建
    if not os.path.exists(RECORD_FILE_ROOT):
        try:
            os.makedirs(RECORD_FILE_ROOT)
            print(f"目录已创建: {RECORD_FILE_ROOT}")
        except Exception as e:
            print(f"创建目录时出错: {e}")
            return

    # 如果未提供 filename，生成基于当前时间的时间戳文件名
    if not filename:
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S%f')
        filename = f"{timestamp}.xlsx"

    # 构建完整的文件路径
    file_path = os.path.join(RECORD_FILE_ROOT, filename)
    file_path = os.path.normpath(file_path).replace('\\', '/')  # 确保路径为Linux风格

    # 创建一个列表来存储每一行的数据
    rows = []

    # 获取所有用户的列表，确保有序（如果需要）
    user_ids = sorted(data.keys())

    for user_id in user_ids:
        functions = data[user_id]
        for func_name, func_data in functions.items():
            record_count = func_data.get('record_count', 0)
            datetime_list = func_data.get('datetime', [])
            # 将日期时间列表转换为逗号分隔的字符串
            datetime_str = ', '.join(datetime_list) if datetime_list else ''
            rows.append({
                '用户ID': user_id,
                '功能': func_name,
                '记录数': record_count,
                '日期时间': datetime_str
            })
        # 在每个用户的数据后插入一个空白行
        rows.append({
            '用户ID': '',  # 空白行的所有字段设为空字符串
            '功能': '',
            '记录数': '',
            '日期时间': ''
        })

    # 创建 DataFrame
    df = pd.DataFrame(rows)

    # 创建新的列列表，包含空白列，确保每个空白列具有唯一的列名
    original_columns = df.columns.tolist()
    new_columns = []
    for col in original_columns:
        new_columns.append(col)
        new_columns.append(f'{col}_blank')  # 使用唯一的列名

    # 初始化一个字典来存储新 DataFrame 的数据
    new_data = {col: [] for col in new_columns}

    # 填充数据
    for index, row in df.iterrows():
        for col in original_columns:
            new_data[col].append(row[col])
            new_data[f'{col}_blank'].append('')  # 空白列填充空字符串

    # 创建新的 DataFrame
    new_df = pd.DataFrame(new_data)

    # 重命名空白列为特定名称（例如：" "）以便在 Excel 中显示为空白
    rename_dict = {f'{col}_blank': ' ' for col in original_columns}
    new_df.rename(columns=rename_dict, inplace=True)

    # 导出到 Excel
    new_df.to_excel(file_path, index=False)

    # 使用 openpyxl 打开 Excel 文件
    wb = load_workbook(file_path)
    ws = wb.active

    # 获取当前的最大列数
    max_col = ws.max_column

    # 插入标题行
    title = "最近一个月用户使用情况"

    # 插入一行到最顶端
    ws.insert_rows(1)

    # 将标题写入第一行的第一个单元格
    ws.cell(row=1, column=1, value=title)

    # 合并标题单元格，跨越所有列
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)

    # 设置标题的样式
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 可选：调整行高以适应标题
    ws.row_dimensions[1].height = 30

    # 保存修改后的 Excel 文件
    wb.save(file_path)
    return file_path, filename

def select_record():
    record = record_count.get_user_record()
    file_path, filename = generate_excel_report(record)
    url = add_url(file_path)
    print(f"url:{url}")
    return url, filename


class SelectRecordHandler(MainHandler):
    @run_on_executor
    def process_async(self):
        return select_record()

    async def post(self):
        url, filename = await self.process_async()
        custom_data = {
            "code": 0,
            "path": url,
            "filename":filename,
            "msg": ''
        }
        self.write(custom_data)









