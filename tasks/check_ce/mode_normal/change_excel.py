import base64
from io import BytesIO
from PIL import Image
import fitz
import openpyxl
from jpype import JClass
from openpyxl.styles import Border, Side
import subprocess
import os
import shutil
import tempfile
import jpype

from .get_table_message import all
from save_filesys_db import save_CE

LIBREOFFICE_PATH = "/usr/bin/soffice"
CODE_SUCCESS = 0
CODE_ERROR = 1
EXCEL_PATH = './assets/excel/temp.xlsx'
IMAGE_PATH = './assets/images/temp.png'
PDF_PATH = './assets/pdf/temp.pdf'
BASE64_PNG = 'data:image/png;base64,'


# 将错误的地方在excel文件框标出
def change_excel(wb, work_table, message_dict):
    """
    该函数主要是为了，呈现错误信息到吉盛标准ce表上
    :param excel: 吉森标准ce表
    :param work_table: ce表中的工作表
    :param message_dict: 错误信息
    :return:
    """
    # 加载工作簿并选择工作表
    # wb = load_workbook(excel)
    sheet = wb[work_table]

    # 用绿色定义边框样式
    # 用绿色定义边框样式
    green_border = Border(
        left=Side(style='thick', color='0000FF'),
        right=Side(style='thick', color='0000FF'),
        top=Side(style='thick', color='0000FF'),
        bottom=Side(style='thick', color='0000FF')
    )

    # 如果message_dict中有'CE-sign'键，处理它的值
    if 'CE-sign' in message_dict:
        ce_values = message_dict['CE-sign']  # 获取'CE-sign'对应的值列表
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value in ce_values:  # 如果单元格的值在ce_values列表中
                    cell.border = green_border  # 更新边框为绿色

    # 遍历工作表中的每一行，更新其他指定的值的边框
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in message_dict and cell.value != 'CE-sign':  # 排除'CE-sign'键
                red_texts_positions = message_dict[cell.value]  # 获取红色文本位置的列表
                red_texts_count = 0  # 找到的红色文本单元格的计数器

                for row_cell in row:
                    # 检查单元格的字体是否为红色
                    if row_cell.font and row_cell.font.color and row_cell.font.color.type == 'rgb' and row_cell.font.color.value == 'FFFF0000':
                        red_texts_count += 1  # 递增计数器
                        # 如果当前计数在位置列表中，则更新边界
                        if red_texts_count in red_texts_positions:
                            row_cell.border = green_border

    # 保存文件
    wb.save(EXCEL_PATH)


# 将excel转化为图片
def excel_to_iamge(excel_path, num):
    output_pdf_path = os.path.splitext(
        os.path.basename(excel_path))[0] + ".pdf"

    # 使用LibreOffice将Excel文件转换为PDF
    try:
        subprocess.run([
            LIBREOFFICE_PATH, "--headless", "--convert-to", "pdf",
            "--outdir", os.getcwd(), excel_path
        ], check=True)
        print(f"PDF successfully created at {output_pdf_path}")
    except subprocess.CalledProcessError as e:
        print(f"Failed to convert Excel to PDF: {e}")
        return

    # 使用fitz (PyMuPDF)将PDF文件的特定页面转换为Base64编码的图片字符串
    doc = fitz.open(output_pdf_path)
    if num < len(doc):
        page = doc.load_page(num)
        pix = page.get_pixmap()
        img_bytes = pix.tobytes("png")
        img_base64 = base64.b64encode(img_bytes).decode("utf-8")
    else:
        img_base64 = None
        print("Specified page number exceeds the PDF's page count.")

    doc.close()
    # 删除生成的PDF文件
    os.remove(output_pdf_path)
    return f"{BASE64_PNG}{img_base64}"


# 将xls文件转化为xlsx文件
def convert_xls_bytes_to_xlsx(file_bytes):
    """
    使用LibreOffice将xls文件的字节流转换为xlsx格式的字节流。
    :param file_bytes: 要转换的xls文件的字节流。
    :return: 转换后的xlsx文件的字节流。
    """
    # 创建一个临时文件来保存字节流
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xls') as tmp_file:
        tmp_file_name = tmp_file.name
        tmp_file.write(file_bytes)
        tmp_file.flush()

    # 获取当前目录作为输出目录
    output_dir = tempfile.mkdtemp()

    try:
        # 构建转换命令
        cmd = [
            LIBREOFFICE_PATH,
            "--headless",
            "--convert-to",
            "xlsx:Calc MS Excel 2007 XML",
            tmp_file_name,
            "--outdir",
            output_dir
        ]
        # 调用命令行执行转换
        subprocess.run(cmd, check=True)
        # 生成的.xlsx文件路径
        generated_file_path = os.path.join(output_dir, os.path.splitext(
            os.path.basename(tmp_file_name))[0] + ".xlsx")
        # 读取生成的.xlsx文件的字节流
        with open(generated_file_path, 'rb') as f:
            xlsx_bytes = f.read()
        return xlsx_bytes

    except subprocess.CalledProcessError as e:
        print(f"转换失败: {e}")
        return None
    except Exception as e:
        print(f"发生错误: {e}")
        return None
    finally:
        # 清理临时文件和目录
        os.remove(tmp_file_name)
        shutil.rmtree(output_dir)


# 判断文件是那种excel文件
def determine_file_type(excel_bytes):
    # 将前几个字节转换为十六进制表示
    file_header = excel_bytes[:8].hex()
    # 判断文件类型
    if file_header.startswith('d0cf11e0a1b11ae1'):
        return 'xls'
    elif file_header.startswith('504b0304'):
        return 'xlsx'
    else:
        return 'Unknown'


def checkTags(username, excel_file, pdf_file, name1, name2, num):
    # 获取文件的目录路径
    directory = os.path.dirname(EXCEL_PATH)
    # 检查这个目录是否存在
    if not os.path.exists(directory):
        # 如果目录不存在，则创建它
        os.makedirs(directory)
        print(f"目录 {directory} 已创建。")
    else:
        print(f"目录 {directory} 已存在。")
    excel_type = determine_file_type(excel_file)
    if excel_type == 'xls':
        excel_file = convert_xls_bytes_to_xlsx(excel_file)
    elif excel_type == 'Unknown':
        code = '该文件不是excel文件'
        return CODE_ERROR, None, code
    doc = fitz.open(stream=BytesIO(pdf_file))
    doc.save(PDF_PATH)
    wb = openpyxl.load_workbook(filename=BytesIO(excel_file))
    sheet_names = wb.sheetnames
    work_table = sheet_names[num]
    if work_table is None:
        sheet_names = wb.sheetnames
        work_table = sheet_names[1]
    print(f"工作表为: {work_table}")
    message_dict = all(wb, work_table, doc, PDF_PATH)
    change_excel(wb, work_table, message_dict)
    image_base64 = excel_to_iamge(EXCEL_PATH, num)
    save_CE(username, doc, EXCEL_PATH, name1, name2,
            work_table, CODE_SUCCESS, image_base64, None)
    os.remove(EXCEL_PATH)
    # os.remove(PDF_PATH)
    doc.close()
    wb.close()
    return CODE_SUCCESS, image_base64, None


# 测试
# def pdf_to_bytes(file_path):
  #  with open(file_path, 'rb') as file:
   #     bytes_content = file.read()
    # return bytes_content
